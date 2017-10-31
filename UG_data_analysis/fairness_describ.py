import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import xlsxwriter



def main_fairness():
    #check the main effect of fairness, whether people accept more fair offer in terms of percentage.
    #5 levels of fairness for the offers (1 = 50/50; 2 = 60/40; 3 = 70/30; 4 = 8/2; 5 = 90/10);
    #fair = 5-5, 6-4;
    #unfair = 7-3, 8-2, 9 -1

    df = pd.read_csv('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/ug_all_task_data.csv', encoding="ISO-8859-1")
    df['fairness']= np.where((df['Fairness_score'] == 1) | (df['Fairness_score'] == 2), 'fair', 'unfair')
    df.to_csv('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/ug_all_task_data.csv', index=False)

    fairCount = df.groupby(['fairness', 'AcceptOffer']).size().unstack(fill_value=0)
    print(fairCount)
    fairCount.to_csv("/Users/kezhang/ownCloud/Suicide_UG/UG_clean_data/fairness.csv")

    #bar graph
    reject_fair = fairCount.iat[0,0]/(fairCount.iat[0,0]+fairCount.iat[0,1])
    reject_unfair = fairCount.iat[1,0]/(fairCount.iat[1,0]+fairCount.iat[1,1])
    y = [reject_fair, reject_unfair]
    print('percentage of fairness trial vs unfair trial',y)
    groups = ['reject_fair', 'reject_unfair']
    x_pos = np.arange(len(groups))
    plt.bar(x_pos,y,align='center')
    plt.xticks(x_pos,groups)

    ax = plt.gca()
    ax.set_ylim([0, 1])
    plt.show()


def punishType():
    ##recode punishing type
    ##recode empty cell as the baseline

    df = pd.read_csv('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/merged_panels/all_task_data.csv',
                     encoding = "ISO-8859-1")
    df['PunishingType'] = np.where(df['PunishingType'].isnull(), 'baseline', df['PunishingType'])
    print(df['PunishingType'])
    df.to_csv('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/ug_all_task_data.csv', index = False)

def byGroup_reappraisal_fair():
    ##How each group accept fairness offer under different reapprisal condition: baseline, punish, empathy.
    ##baseline condition and experimental condition are all combine

    df = pd.read_csv('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\ug_all_task_data.csv', encoding = "ISO-8859-1")
    ct_group5_reappraisal = pd.crosstab([df.group5, df.ReappraisalDirection, df.fairness], df.AcceptOffer, normalize='index')
    ct_group4_reappraisal = pd.crosstab([df.group4, df.ReappraisalDirection, df.fairness], df.AcceptOffer, normalize='index')

    workbook = xlsxwriter.Workbook('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\fair_reappraisal.xlsx')
    workbook.close()

    book = load_workbook('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\fair_reappraisal.xlsx')
    writer = pd.ExcelWriter('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\fair_reappraisal.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    ct_group5_reappraisal.to_excel(writer, 'group5_reappraisal_percent')
    ct_group4_reappraisal.to_excel(writer, 'group4_reappraisal_percent')
    writer.save()
    writer.close()

def fairness():
    #check the main effect of fairness, whether people accept more fair offer in terms of percentage.
    #5 levels of fairness for the offers (1 = 50/50; 2 = 60/40; 3 = 70/30; 4 = 8/2; 5 = 90/10);
    #fair = 5-5, 6-4;
    #unfair = 7-3, 8-2, 9 -1

    df = pd.read_csv('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/ug_all_task_data.csv', encoding="ISO-8859-1")
    df['fairness']= np.where((df['Fairness_score'] == 1) | (df['Fairness_score'] == 2), 'fair', 'unfair')
    df.to_csv('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/ug_all_task_data.csv', index=False)

def byGroup_baseline_fairness():
    # only have Baseline condition
    # crosstab between fairness and groups
    # two-way chi-square
    all_data = pd.read_csv('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/ug_all_task_data.csv', encoding="ISO-8859-1")
    baseline = all_data[all_data.ReappraisalDirection == 'baseline']

    ct_group5 = pd.crosstab([baseline.group5, baseline.fairness], baseline.AcceptOffer)
    ct_group4 = pd.crosstab([baseline.group4, baseline.fairness], baseline.AcceptOffer)

    # ct_group5_p = ct_group5/ct_group5.groupby(level = 0, axis=1).sum()
    # ct_group4_p = ct_group4/ct_group4.groupby(level = 0, axis=1).sum()

    ct_group5_p = pd.crosstab([baseline.group5, baseline.fairness], baseline.AcceptOffer, normalize='index')
    ct_group4_p = pd.crosstab([baseline.group4, baseline.fairness], baseline.AcceptOffer, normalize='index')

    book = load_workbook('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/chisquare_crosstab.xlsx')
    writer = pd.ExcelWriter('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/chisquare_crosstab.xlsx',
                            engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    ct_group5.to_excel(writer, "baseline_group4_count")
    ct_group4.to_excel(writer, "baseline_group5_count")
    ct_group4_p.to_excel(writer, 'baseline_group4')
    ct_group5_p.to_excel(writer, 'baseline_group5')
    writer.save()

def stack_accept_baseline():
    # arrange stack size and acceptance rate in baseline data
    all_data = pd.read_csv('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\ug_all_task_data.csv', encoding="ISO-8859-1")

    baseline = all_data[all_data.ReappraisalDirection == 'baseline']
    ct1 = pd.crosstab([baseline.ID, baseline.TotalAmountAtStake], baseline.AcceptOffer, margins=True)
    ct2 = pd.crosstab([baseline.ID, baseline.TotalAmountAtStake], baseline.AcceptOffer, normalize='index')
    ct3 = pd.crosstab([baseline.TotalAmountAtStake, baseline.group5], baseline.AcceptOffer, margins=True)
    ct4 = pd.crosstab([baseline.TotalAmountAtStake, baseline.group5], baseline.AcceptOffer, normalize='index')
    ct5 = pd.crosstab([baseline.PlayerProposedAmount, baseline.group5], baseline.AcceptOffer, margins=True)
    ct6 = pd.crosstab([baseline.PlayerProposedAmount, baseline.group5], baseline.AcceptOffer, normalize='index')
    ct7 = pd.crosstab([baseline.Fairness_score, baseline.group5], baseline.AcceptOffer, margins=True)
    ct8 = pd.crosstab([baseline.Fairness_score, baseline.group5], baseline.AcceptOffer, normalize='index')
    ct13 = pd.crosstab([baseline.OpponentProposedAmount, baseline.group5], baseline.AcceptOffer, margins=True)
    ct14 = pd.crosstab([baseline.OpponentProposedAmount, baseline.group5], baseline.AcceptOffer, normalize='index')
    print(baseline.PlayerProposedAmount.unique())
    print(ct5)

    workbook = xlsxwriter.Workbook('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\baseline_stackSize.xlsx')
    workbook.close()
    book = load_workbook('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\stackSize_baseline.xlsx')
    writer = pd.ExcelWriter('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\stackSize_baseline.xlsx', engine = 'openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    ct1.to_excel(writer, 'totalSizeByID_count')
    ct2.to_excel(writer, 'totalSizeByID_percent')
    ct3.to_excel(writer, 'totalSizeByGroup_count')
    ct4.to_excel(writer, 'totalSizeByGroup_percent')

    ct5.to_excel(writer, 'proposedSizeByGroup_count')
    ct6.to_excel(writer, 'proposedSizeByGroup_percent')
    ct7.to_excel(writer, 'fairscoreByGroup_count')
    ct8.to_excel(writer, 'fairscoreByGroup_percent')
    ct13.to_excel(writer, 'oppoProposedSize_count')
    ct14.to_excel(writer, 'oppoProposedSize_percent')

    writer.save()


def stack_accept_framing():
    all_data = pd.read_csv('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\ug_all_task_data.csv', encoding="ISO-8859-1")
    punish = all_data[all_data.ReappraisalDirection == 'punish']
    empathy = all_data[all_data.ReappraisalDirection == 'empathy']

    ct1 = pd.crosstab([punish.TotalAmountAtStake, punish.group5], punish.AcceptOffer, margins=True)
    ct2 = pd.crosstab([punish.TotalAmountAtStake, punish.group5], punish.AcceptOffer, normalize='index')
    ct3 = pd.crosstab([punish.PlayerProposedAmount, punish.group5], punish.AcceptOffer, margins=True)
    ct4 = pd.crosstab([punish.PlayerProposedAmount, punish.group5], punish.AcceptOffer, normalize='index')
    ct5 = pd.crosstab([punish.Fairness_score, punish.group5], punish.AcceptOffer, margins=True)
    ct6 = pd.crosstab([punish.Fairness_score, punish.group5], punish.AcceptOffer, normalize='index')
    ct13 = pd.crosstab([punish.OpponentProposedAmount, punish.group5], punish.AcceptOffer, margins=True)
    ct14 = pd.crosstab([punish.OpponentProposedAmount, punish.group5], punish.AcceptOffer, normalize='index')


    workbook = xlsxwriter.Workbook('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/punish_stackSize.xlsx')
    workbook.close()
    book = load_workbook('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\stackSize_punish.xlsx')
    writer = pd.ExcelWriter('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\stackSize_punish.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    ct1.to_excel(writer, 'totalSizeByGroup_count')
    ct2.to_excel(writer, 'totalSizeByGroup_percent')
    ct3.to_excel(writer, 'proposedSizeByGroup_count')
    ct4.to_excel(writer, 'proposedSizeByGroup_percent')
    ct5.to_excel(writer, 'fairscoreByGroup_count')
    ct6.to_excel(writer, 'fairscoreByGroup_percent')
    ct13.to_excel(writer, 'oppoProposedSize_count')
    ct14.to_excel(writer, 'oppoProposedSize_percent')
    writer.save()
    writer.close()

    ct7 = pd.crosstab([empathy.TotalAmountAtStake, empathy.group5], empathy.AcceptOffer, margins=True)
    ct8 = pd.crosstab([empathy.TotalAmountAtStake, empathy.group5], empathy.AcceptOffer, normalize='index')
    ct9 = pd.crosstab([empathy.PlayerProposedAmount, empathy.group5], empathy.AcceptOffer, margins=True)
    ct10 = pd.crosstab([empathy.PlayerProposedAmount, empathy.group5], empathy.AcceptOffer, normalize='index')
    ct11 = pd.crosstab([empathy.Fairness_score, empathy.group5], empathy.AcceptOffer, margins=True)
    ct12 = pd.crosstab([empathy.Fairness_score, empathy.group5], empathy.AcceptOffer, normalize='index')
    ct15 = pd.crosstab([empathy.OpponentProposedAmount, empathy.group5], empathy.AcceptOffer, margins=True)
    ct16 = pd.crosstab([empathy.OpponentProposedAmount, empathy.group5], empathy.AcceptOffer, normalize='index')

    # workbook = xlsxwriter.Workbook('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/empathy_stackSize.xlsx')
    # workbook.close()
    book = load_workbook('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\stackSize_empathy.xlsx')
    writer = pd.ExcelWriter('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\stackSize_empathy.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    ct7.to_excel(writer, 'totalSizeByGroup_count')
    ct8.to_excel(writer, 'totalSizeByGroup_percent')
    ct9.to_excel(writer, 'proposedSizeByGroup_count')
    ct10.to_excel(writer, 'proposedSizeByGroup_percent')
    ct11.to_excel(writer, 'fairscoreByGroup_count')
    ct12.to_excel(writer, 'fairscoreByGroup_percent')
    ct15.to_excel(writer, 'oppoProposedSize_count')
    ct16.to_excel(writer, 'oppoProposedSize_percent')
    writer.save()
    writer.close()

def stake_accept_reappraisal_control():
    # acceptance rate and stake size of HC
    # acceptance rate with stake size (player proposed size), social framing
    # acceptance rate with fairness (binary), social framing
    # acceptance rate with fairness (by 5 level proportions), social framing

    controls = pd.read_excel('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/controls_data.xlsx', encoding="ISO-8859-1")
    ct1 = pd.crosstab(controls.PlayerProposedAmount, controls.AcceptOffer, margins=True)
    ct2 = pd.crosstab(controls.PlayerProposedAmount, controls.AcceptOffer, normalize='index')
    ct3 = pd.crosstab(controls.Fairness_score, controls.AcceptOffer, margins=True)
    ct4 = pd.crosstab(controls.Fairness_score, controls.AcceptOffer, normalize='index')
    ct5 = pd.crosstab([controls.PlayerProposedAmount, controls.ReappraisalDirection], controls.AcceptOffer, margins=True)
    ct6 = pd.crosstab([controls.PlayerProposedAmount, controls.ReappraisalDirection], controls.AcceptOffer, normalize='index')
    ct7 = pd.crosstab([controls.fairness, controls.ReappraisalDirection], controls.AcceptOffer, margins=True)
    ct8 = pd.crosstab([controls.fairness, controls.ReappraisalDirection], controls.AcceptOffer, normalize='index')
    ct9 = pd.crosstab([controls.Fairness_score, controls.ReappraisalDirection], controls.AcceptOffer, margins=True)
    ct10 = pd.crosstab([controls.Fairness_score, controls.ReappraisalDirection], controls.AcceptOffer, normalize='index')

    workbook = xlsxwriter.Workbook('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/stackSize_controls.xlsx')
    workbook.close()
    book = load_workbook('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/stackSize_controls.xlsx')
    writer = pd.ExcelWriter('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/stackSize_controls.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    ct1.to_excel(writer, 'proposedSize_count')
    ct2.to_excel(writer, 'proposedSize_percent')
    ct3.to_excel(writer, 'fairscore_count')
    ct4.to_excel(writer, 'fairscore_percent')
    ct5.to_excel(writer, 'proposedSizeFraming_count')
    ct6.to_excel(writer, 'proposedSizeFraming_percent')
    ct7.to_excel(writer, 'fairFraming_count')
    ct8.to_excel(writer, 'fairFraming_percent')
    ct9.to_excel(writer, 'fairscoreFraming_count')
    ct10.to_excel(writer, 'fairscoreFraming_percent')
    writer.save()
    writer.close()

def byGroup_reappraisal_fairness():
    all_data = pd.read_csv('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/ug_all_task_data.csv', encoding="ISO-8859-1")
    reappra = all_data[all_data.ReappraisalDirection != 'baseline']

    ct_group5 = pd.crosstab(reappra.group5, [reappra.fairness, reappra.ReappraisalDirection, reappra.AcceptOffer], margins=True)
    ct_group4 = pd.crosstab(reappra.group4, [reappra.fairness, reappra.ReappraisalDirection, reappra.AcceptOffer], margins=True)
    # ct_group5_p = pd.crosstab([reappra.group5, reappra.fairness, reappra.ReappraisalDirection], reappra.AcceptOffer, normalize='index')
    # ct_group4_p = pd.crosstab([reappra.group4, reappra.fairness, reappra.ReappraisalDirection], reappra.AcceptOffer, normalize='index')


    book = load_workbook('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/chisquare_crosstab.xlsx')
    writer = pd.ExcelWriter('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/chisquare_crosstab.xlsx',
                            engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    ct_group4.to_excel(writer, 'group4_3way_count')
    # ct_group5.to_excel(writer, 'group5_3way_graph')
    # ct_group5_p.to_excel(writer, 'group5_3way')
    # ct_group4_p.to_excel(writer, 'group4_3way')
    writer.save()


def tryversion():
    df = pd.read_excel('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\try.xlsx')
    ct = pd.crosstab(df.group, [df.position, df.offer], margins=True)
    print(ct)

