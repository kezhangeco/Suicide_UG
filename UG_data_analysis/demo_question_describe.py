import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from scipy import stats
from statsmodels.stats.multicomp import (pairwise_tukeyhsd, MultiComparison)


def byGroup():
    # new var identifies the subject's group: control, depressed_control, ideator, AttempterHL, AttempterLL
    # group 5: split attempter into HL and LL
    # group 4: does not split attempters
    df = pd.read_csv('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/merged_panels/all_task_data.csv',  encoding="ISO-8859-1",
                     dtype={'StakeImg': object, 'ReappraisalText': object, 'ReappraisalDirection': object,
                            'PunishingType': object})


    df["group5"] = np.where(df['PATTYPE'] == 'CONTROL', 'control', np.where(df['PATTYPE'] == 'DEPRESSION', 'depression',
                                                                           np.where(df['COMMENT'] == 'IDEATOR', 'ideator',
                                                                                    np.where(df['COMMENT'] == 'IDEATOR-ATTEMPTER',
                                                                                             np.where(df['MAXLETHALITY'] < 4, 'AttempterLL', 'AttempterHL'),
                                                                                             np.where(df['COMMENT'] == 'ATTEMPTER',
                                                                                                      np.where(df['MAXLETHALITY'] < 4, 'AttempterLL', 'AttempterHL'), 'NA')))))

    df['group4'] = np.where((df['group5'] == 'AttempterLL') | (df['group5'] == 'AttempterHL'), 'attempter', df['group5'])
    df.to_csv('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/merged_panels/ug_all_task_data.csv', encoding = "ISO-8859-1", index = False)


def demo_description():
    # add 'group' variable into demographic data: 5 levels group and 4 levels group. HH and LL attempters
    # describe demographic info, mean, median, sd...

    df1 = pd.read_csv('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/ug_demog.csv')
    ID120 = pd.read_csv("/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/wrongData/ug_all_task_data.csv",
                        encoding="ISO-8859-1")

    book = load_workbook('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/summary.xlsx')
    writer = pd.ExcelWriter('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/summary.xlsx', engine='openpyxl')
    writer.book = book

    group5_demo = df1.groupby(['group5'])['PROTECT2 AGE', 'MARITAL TEXT', 'GENDER TEXT',
                                          'EDUCATION', 'RACE TEXT'].describe(percentiles=None, include = 'all')
    group4_demo = df1.groupby(['group4'])['PROTECT2 AGE', 'MARITAL TEXT', 'GENDER TEXT',
                                          'EDUCATION'].describe(percentiles=None, include = 'all')



    writer = pd.ExcelWriter("/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/summary.xlsx")
    group5_demo.to_excel(writer, 'group5_demo')
    group4_demo.to_excel(writer, 'group4_demo')
    writer.save()
    writer.close()
    idls = np.unique(df1['ID'])
    # print(idls)
    # print(len(idls))

    HL = np.unique(df1[df1['group5'] == 'AttempterHL']['ID'])
    LL = np.unique(df1[df1['group5'] == 'AttempterLL']['ID'])
    c = np.unique(df1[df1['group5'] == 'control']['ID'])
    d = np.unique(df1[df1['group5'] == 'depression']['ID'])
    i = np.unique(df1[df1['group5'] == 'ideator']['ID'])
    # print(bool(set(HL) & set(LL) & set(c) & set(d) & set(i)))
    # print(HL, len(HL))
    # print(LL, len(LL))
    # print(c, len(c))
    # print(d, len(d))
    # print(i, len(i))

    dID = df1[df1['group5'] == 'depression']['ID']
    # print(list(dID), len(dID))
    # print(len(np.unique(dID)))
    # print(list(set(dID)- (set(d))))
    print(dID[dID.duplicated()])



def questionnaires_description():
    # describe questionnaires information, such as depression, MMSE and others

    df = pd.read_excel('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/ug_questionnaire.xlsx')
    book = load_workbook("/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/demo_summary.xlsx")
    writer = pd.ExcelWriter("/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/demo_summary.xlsx", engine = 'openpyxl')
    writer.book = book

    # question_bygroup5 = df.groupby('group5')['DEP ONSET AGE', 'HOUSEHOLD INCOME', 'HRSD NO SUI', 'MMSE TOTAL',
    # 'PER CAPITA INCOME', 'SSI BL CURRENT', 'DRS', 'MAX LETHALITY', 'SIS ML PLAN', 'SIS ML TOTAL', 'Anxiety Current',
    #                                          'Anxiety Lifetime', 'Substance Current', 'Substance Lifetime',
    #                                          'EXIT'].describe(include='all')
    # question_bygroup4 = df.groupby('group4')['DEP ONSET AGE', 'HOUSEHOLD INCOME', 'HRSD NO SUI', 'MMSE TOTAL',
    # 'PER CAPITA INCOME', 'SSI BL CURRENT', 'DRS', 'MAX LETHALITY', 'SIS ML PLAN', 'SIS ML TOTAL', 'Anxiety Current',
    #                                          'Anxiety Lifetime', 'Substance Current', 'Substance Lifetime',
    #                                          'EXIT'].describe(include='all')

    question_bygroup5 = df.groupby('group5')['MAX LETHALITY', 'SIS ML PLAN', 'SIS ML TOTAL', 'Anxiety Current',
                                             'Anxiety Lifetime', 'Substance Current', 'Substance Lifetime',
                                             'EXIT'].describe(include='all')
    question_bygroup4 = df.groupby('group4')['MAX LETHALITY', 'SIS ML PLAN', 'SIS ML TOTAL', 'Anxiety Current',
                                             'Anxiety Lifetime', 'Substance Current', 'Substance Lifetime',
                                             'EXIT'].describe(include='all')

    question_bygroup5.to_excel(writer, 'group5_questionnaires')
    question_bygroup4.to_excel(writer, 'group4_questionnaired')
    writer.save()
    writer.close()


def demo_summ_clean():
    summ5 = pd.read_excel('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/UG_summary.xlsx', sheetname='group5')
    summ4 = pd.read_excel('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/UG_summary.xlsx', sheetname='group4')

    rows = ['min', 'max', '25%', '50%', '75%', 'max', 'unique', 'top', 'freq']

    summ_g5 = ~summ5.stats.isin(rows)
    summ_g5 = summ5[summ_g5]

    summ_g4 = ~summ4.stats.isin(rows)
    summ_g4 = summ4[summ_g4]


    book = load_workbook('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/UG_summary_py.xlsx')
    writer = pd.ExcelWriter('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/UG_summary_py.xlsx', engine='openpyxl')
    writer.book = book

    summ_g5.to_excel(writer, 'group5', index = False)
    summ_g4.to_excel(writer, 'group4', index = False)
    writer.save()
    writer.close()


def summ_gender():
    summ = pd.read_csv('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/ug_demog.csv')
    summ = summ[['ID', 'GENDERTEXT', 'RACETEXT', 'MARITALTEXT', 'group5', 'group4']]
    summ.to_excel('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/ug_gender.xlsx')

def compareMean():
    # df = pd.read_excel('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/ug_questionnaire.xlsx')
    # df1 = pd.read_excel('/Users/kezhang/ownCloud/Suicide_UG/UG_clean_updated/ug_demog.xlsx')
    df1 = pd.read_excel('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\ug_demog.xlsx')
    df = pd.read_excel('C:\\Users\\ke\\ownCloud\\Suicide_UG\\UG_clean_updated\\ug_questionnaire.xlsx')

    df1[['group5'], ['group4']] = df1[['group5'], ['group4']].astype(str)
    print('dtype ', df1['group5'].dtype)

    demo_ls = ['PROTECT2AGE', 'MARITALTEXT', 'GENDERTEXT', 'EDUCATION', 'RACETEXT']
    question_ls = ['DEP ONSET AGE', 'HOUSEHOLD INCOME', 'HRSD NO SUI', 'MMSE TOTAL', 'PER CAPITA INCOME',
                   'SSI BL CURRENT', 'DRS']

    # demo stats
    attempter = df1[df1['group4'] == 'attempter'][demo_ls]
    attempterHL = df1[df1['group5'] == 'AttempterHL'][demo_ls]
    attempterLL = df1[df1['group5'] == 'AttempterLL'][demo_ls]
    ideator = df1[df1['group4'] == 'ideator'][demo_ls]
    control = df1[df1['group4'] == 'control'][demo_ls]
    depression = df1[df1['group4'] == 'depression'][demo_ls]



    ## demo ANOVAs
    demo_anova_ls = ['PROTECT2AGE', 'EDUCATION']
    for i in demo_anova_ls:
        f, p = stats.f_oneway(attempterLL[i], attempterHL[i], ideator[i], control[i], depression[i])
        print('5 groups', i, 'f value: ', f, 'p value: ', p)

    for j in demo_anova_ls:
        f, p = stats.f_oneway(attempter[j], ideator[j], control[j], depression[j])
        print('4 groups', j, 'f value: ', f, 'p value: ', p)

    tukey = pairwise_tukeyhsd(df1['group5'], df1['PROTECT2AGE'])
    print(tukey[0])


    # questionnaires stats
    attempter = df[df['group4'] == 'attempter'][question_ls]
    attempterHL = df[df['group5'] == 'AttempterHL'][question_ls]
    attempterLL = df[df['group5'] == 'AttempterLL'][question_ls]
    ideator = df[df['group4'] == 'ideator'][question_ls]
    control = df[df['group4'] == 'control'][question_ls]
    depression = df[df['group4'] == 'depression'][question_ls]

    ## questionnaires ANOVAs count for all groups, not exclude controls
    question_all_ls = ['HOUSEHOLD INCOME', 'MMSE TOTAL', 'PER CAPITA INCOME', 'DRS']
    question_exclHC_ls = ['DEP ONSET AGE', 'HRSD NO SUI', 'SSI BL CURRENT']


    for k in question_all_ls:
        f, p = stats.f_oneway(attempterLL[k], attempterHL[k], ideator[k], control[k], depression[k])
        print('5 groups', k, 'f value: ', f, 'p value: ', p)

    for l in question_all_ls:
        f, p = stats.f_oneway(attempter[l], ideator[l], control[l], depression[l])
        print('4 groups', l, 'f value: ', f, 'p value: ', p)

    ## questionnaires ANOVAs count for groups exclude controls
    for m in question_exclHC_ls:
        f, p = stats.f_oneway(attempterLL[m], attempterHL[m], ideator[m], depression[m])
        print('two attempters No control', m, 'f value: ', f, 'p value: ', p)

    for n in question_exclHC_ls:
        f, p = stats.f_oneway(attempter[n], ideator[n], depression[n])
        print('One attempter No control', n, 'f value: ', f, 'p value: ', p)




